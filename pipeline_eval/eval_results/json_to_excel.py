import os
import sys
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

if sys.platform.startswith('win'):
    os.system('chcp 65001 > nul')
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

def process_eval_json(json_path: str):
    """
    Đọc 1 file JSON và xuất ra 1 file Excel.
    Các chỉ số summary sẽ được thêm vào thành các cột mới (mỗi metric 1 cột) bên cạnh các cột chi tiết.
    """
    print(f"Đang xử lý: {json_path}")
    
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
        
    summary = data.get("summary", {})
    results = data.get("results", [])
    
    if not results:
        print(f"Không có dữ liệu results trong {json_path}")
        return

    # 1. Chuyển list results thành DataFrame với các cột cũ
    df_results = pd.DataFrame(results)
    
    # Format lại cột list thành string cho Excel
    if 'evidence' in df_results.columns:
        df_results['evidence'] = df_results['evidence'].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)
    if 'bm25_top3_ids' in df_results.columns:
        df_results['bm25_top3_ids'] = df_results['bm25_top3_ids'].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)
        
    # 2. Thêm các metric summary thành các cột mới (mỗi metric 1 cột)
    for key, val in summary.items():
        # Đổi tên cột cho đẹp
        col_name = f"Summary_{key}"
        # Gán giá trị này cho TẤT CẢ các dòng (hoặc chỉ dòng đầu tiên tùy ý, ở đây gán hết để filter dễ)
        df_results[col_name] = val
        
    # Tạo tên file Excel từ file JSON
    base_name = os.path.basename(json_path).replace('.json', '.xlsx')
    excel_path = os.path.join(os.path.dirname(json_path), base_name)
    
    # 3. Ghi DataFrame vào Excel
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')
    df_results.to_excel(writer, sheet_name='Evaluation Results', index=False)
    
    # Tự động điều chỉnh độ rộng cột
    worksheet = writer.sheets['Evaluation Results']
    for idx, col in enumerate(df_results.columns):
        max_len = max(df_results[col].astype(str).map(len).max(), len(col)) + 2
        worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_len, 50)

    writer.close()
    print(f"Đã xuất thành công: {excel_path}\n")

if __name__ == "__main__":
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    json_files = [f for f in os.listdir(current_dir) if f.endswith('.json')]
    
    if not json_files:
        print("Không tìm thấy file JSON nào trong thư mục eval_results!")
    else:
        for json_file in json_files:
            process_eval_json(os.path.join(current_dir, json_file))
